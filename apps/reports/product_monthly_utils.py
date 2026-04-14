"""
Extra report generators for top-products and monthly-revenue reports.
Imported by utils.py via star-import shim.
"""
from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BRAND_COLOR  = colors.HexColor('#C00000')
HEADER_BG    = colors.HexColor('#0f3460')
ACCENT_COLOR = colors.HexColor('#F5B400')
ALT_ROW      = colors.HexColor('#f0f4ff')
_thin   = Side(style='thin', color='CCCCCC')
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr_style(styles):
    return ParagraphStyle('H', parent=styles['Heading1'], fontSize=17,
                          textColor=BRAND_COLOR, alignment=TA_CENTER, spaceAfter=4)


def _sub_style(styles):
    return ParagraphStyle('S', parent=styles['Normal'], fontSize=9,
                          textColor=colors.HexColor('#666666'), alignment=TA_CENTER, spaceAfter=12)


def _tbl_style(total=False):
    base = [
        ('BACKGROUND',  (0, 0), (-1, 0),  HEADER_BG),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  10),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',    (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ALT_ROW]),
        ('GRID',        (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
    ]
    if total:
        base += [
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR',  (0, -1), (-1, -1), ACCENT_COLOR),
            ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]
    return TableStyle(base)


def _xl_setup(ws, title, subtitle, headers):
    n = len(headers)
    end_col = chr(64 + n)
    ws.merge_cells(f'A1:{end_col}1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=15, color='C00000')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'A2:{end_col}2')
    ws['A2'] = subtitle
    ws['A2'].font = Font(size=9, color='666666', italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    hfill = PatternFill(start_color='0f3460', end_color='0f3460', fill_type='solid')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal='center')
        c.border = _border


def _xl_row(ws, row_idx, values, alt=False):
    fill = PatternFill(start_color='f0f4ff', end_color='f0f4ff', fill_type='solid') if alt else None
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=col, value=v)
        c.alignment = Alignment(horizontal='center')
        c.border = _border
        if fill:
            c.fill = fill


# ── Top Products ──────────────────────────────────────────────────────────────

def generate_products_report_pdf(products_data, period_label):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=0.6*inch, rightMargin=0.6*inch,
                            topMargin=0.6*inch, bottomMargin=0.6*inch)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("ColdSync Pro — Top Products Report", _hdr_style(styles)),
        Paragraph(f"{period_label}  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", _sub_style(styles)),
    ]
    headers = ['#', 'Product', 'Brand', 'Size', 'Bottles Sold', 'Orders', 'Revenue (₹)']
    rows = [headers]
    grand = 0
    for i, p in enumerate(products_data, 1):
        rows.append([str(i), p['product_name'][:22], p['brand'], p['bottle_size'],
                     str(p['total_bottles']), str(p['order_count']), f"₹{p['total_revenue']:,.0f}"])
        grand += p['total_revenue']
    rows.append(['', 'TOTAL', '', '', '', '', f"₹{grand:,.0f}"])
    t = Table(rows, colWidths=[0.35*inch, 1.9*inch, 1.1*inch, 0.7*inch, 1*inch, 0.7*inch, 1.1*inch])
    t.setStyle(_tbl_style(total=True))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_products_report_excel(products_data, period_label):
    buffer = BytesIO()
    wb = Workbook(); ws = wb.active; ws.title = 'Top Products'
    headers = ['#', 'Product', 'Brand', 'Size', 'Bottles Sold', 'Orders', 'Revenue (₹)']
    _xl_setup(ws, 'ColdSync Pro — Top Products Report', period_label, headers)
    grand = 0
    for i, p in enumerate(products_data, 1):
        _xl_row(ws, i + 4, [i, p['product_name'], p['brand'], p['bottle_size'],
                             p['total_bottles'], p['order_count'], p['total_revenue']], alt=(i % 2 == 0))
        grand += p['total_revenue']
    tr = len(products_data) + 5
    ws.cell(tr, 1, 'TOTAL').font = Font(bold=True)
    ws.cell(tr, 7, grand).font = Font(bold=True, color='C00000')
    for col, w in zip('ABCDEFG', [5, 26, 14, 10, 14, 10, 16]):
        ws.column_dimensions[col].width = w
    wb.save(buffer); buffer.seek(0)
    return buffer


# ── Monthly Revenue ───────────────────────────────────────────────────────────

def generate_monthly_report_pdf(months_data):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=0.6*inch, rightMargin=0.6*inch,
                            topMargin=0.6*inch, bottomMargin=0.6*inch)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("ColdSync Pro — Monthly Revenue Report", _hdr_style(styles)),
        Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", _sub_style(styles)),
    ]
    headers = ['Month', 'Revenue (₹)', 'Orders', 'Avg Order (₹)']
    rows = [headers]
    grand = 0
    for m in months_data:
        rows.append([m['month'], f"₹{m['revenue']:,.0f}", str(m['order_count']), f"₹{m['avg_order']:,.0f}"])
        grand += m['revenue']
    rows.append(['TOTAL', f"₹{grand:,.0f}", '', ''])
    t = Table(rows, colWidths=[1.5*inch, 1.8*inch, 1.2*inch, 1.8*inch])
    t.setStyle(_tbl_style(total=True))
    elements.append(t)
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_monthly_report_excel(months_data):
    buffer = BytesIO()
    wb = Workbook(); ws = wb.active; ws.title = 'Monthly Revenue'
    headers = ['Month', 'Revenue (₹)', 'Orders', 'Avg Order (₹)']
    _xl_setup(ws, 'ColdSync Pro — Monthly Revenue Report',
              f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}', headers)
    grand = 0
    for i, m in enumerate(months_data, 1):
        _xl_row(ws, i + 4, [m['month'], m['revenue'], m['order_count'], m['avg_order']], alt=(i % 2 == 0))
        grand += m['revenue']
    tr = len(months_data) + 5
    ws.cell(tr, 1, 'TOTAL').font = Font(bold=True)
    ws.cell(tr, 2, grand).font = Font(bold=True, color='C00000')
    for col, w in zip('ABCD', [16, 16, 12, 16]):
        ws.column_dimensions[col].width = w
    wb.save(buffer); buffer.seek(0)
    return buffer
