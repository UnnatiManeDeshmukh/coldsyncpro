"""
Report generation utilities — PDF (ReportLab) + Excel (openpyxl)
"""
import io
from django.utils import timezone


# ── helpers ──────────────────────────────────────────────────────────────────

def _header(pdf, title, subtitle=''):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    styles = getSampleStyleSheet()
    pdf.setFillColor(colors.HexColor('#c0392b'))
    pdf.rect(0, A4[1] - 80, A4[0], 80, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 20)
    pdf.drawString(40, A4[1] - 45, 'ColdSync Pro — Shree Ganesh Agency')
    pdf.setFont('Helvetica', 13)
    pdf.drawString(40, A4[1] - 65, title)
    if subtitle:
        pdf.setFont('Helvetica', 10)
        pdf.drawString(40, A4[1] - 78, subtitle)


def _footer(pdf, page_num=1):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    pdf.setFillColor(colors.HexColor('#ecf0f1'))
    pdf.rect(0, 0, A4[0], 30, fill=True, stroke=False)
    pdf.setFillColor(colors.HexColor('#7f8c8d'))
    pdf.setFont('Helvetica', 8)
    pdf.drawString(40, 10, f'Generated: {timezone.now().strftime("%d %b %Y %I:%M %p")}  |  Page {page_num}')
    pdf.drawRightString(A4[0] - 40, 10, 'ColdSync Pro © 2026')


# ── Stock Report ──────────────────────────────────────────────────────────────

def generate_stock_report_pdf(stock_data):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors

    buf = io.BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    W, H = A4

    _header(pdf, 'Inventory / Stock Report', f'Total Products: {len(stock_data)}')

    # Table header
    y = H - 110
    cols = [40, 180, 300, 390, 460, 530]
    headers = ['#', 'Product', 'Brand', 'Warehouse', 'Crates', 'Bottles']
    pdf.setFillColor(colors.HexColor('#2c3e50'))
    pdf.rect(30, y - 5, W - 60, 20, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 9)
    for i, h in enumerate(headers):
        pdf.drawString(cols[i], y + 3, h)

    pdf.setFont('Helvetica', 9)
    y -= 20
    for idx, row in enumerate(stock_data):
        if y < 60:
            _footer(pdf)
            pdf.showPage()
            y = H - 50
        bg = colors.HexColor('#f8f9fa') if idx % 2 == 0 else colors.white
        pdf.setFillColor(bg)
        pdf.rect(30, y - 5, W - 60, 18, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor('#2c3e50'))
        vals = [str(idx + 1), str(row['product'])[:22], str(row['brand'])[:15],
                str(row['warehouse'])[:15], str(row['total_crates']), str(row['total_bottles'])]
        for i, v in enumerate(vals):
            pdf.drawString(cols[i], y + 2, v)
        y -= 18

    _footer(pdf)
    pdf.save()
    buf.seek(0)
    return buf


def generate_stock_report_excel(stock_data):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stock Report'

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'ColdSync Pro — Inventory / Stock Report'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='C0392B')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Generated: {timezone.now().strftime("%d %b %Y %I:%M %p")}'
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['#', 'Product', 'Brand', 'Warehouse', 'Total Crates', 'Total Bottles']
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2C3E50')
        cell.alignment = Alignment(horizontal='center')

    for idx, row in enumerate(stock_data, 1):
        ws.append([idx, row['product'], row['brand'], row['warehouse'],
                   row['total_crates'], row['total_bottles']])
        if idx % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill('solid', fgColor='F8F9FA')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Sales Report ──────────────────────────────────────────────────────────────

def generate_sales_report_pdf(sales_data, month_label):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors

    buf = io.BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    total = sum(r['total_amount'] for r in sales_data)

    _header(pdf, f'Sales Report — {month_label}',
            f'Total Orders: {len(sales_data)}  |  Total Revenue: ₹{total:,.2f}')

    y = H - 110
    cols = [40, 80, 220, 320, 400, 490]
    headers = ['#', 'Customer', 'Date', 'Amount (₹)', 'Payment', 'Delivery']
    pdf.setFillColor(colors.HexColor('#2c3e50'))
    pdf.rect(30, y - 5, W - 60, 20, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 9)
    for i, h in enumerate(headers):
        pdf.drawString(cols[i], y + 3, h)

    pdf.setFont('Helvetica', 8)
    y -= 20
    for idx, row in enumerate(sales_data):
        if y < 60:
            _footer(pdf)
            pdf.showPage()
            y = H - 50
        bg = colors.HexColor('#f8f9fa') if idx % 2 == 0 else colors.white
        pdf.setFillColor(bg)
        pdf.rect(30, y - 5, W - 60, 18, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor('#2c3e50'))
        date_str = row['date'].strftime('%d %b %Y') if hasattr(row['date'], 'strftime') else str(row['date'])[:10]
        vals = [str(idx + 1), str(row['customer'])[:20], date_str,
                f"₹{row['total_amount']:,.0f}", row['payment_status'], row['delivery_status'][:12]]
        for i, v in enumerate(vals):
            pdf.drawString(cols[i], y + 2, v)
        y -= 18

    _footer(pdf)
    pdf.save()
    buf.seek(0)
    return buf


def generate_sales_report_excel(sales_data, month_label):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Report'

    ws.merge_cells('A1:F1')
    ws['A1'] = f'ColdSync Pro — Sales Report — {month_label}'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='C0392B')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Generated: {timezone.now().strftime("%d %b %Y %I:%M %p")}'
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])
    ws.append(['#', 'Customer', 'Date', 'Amount (₹)', 'Payment Status', 'Delivery Status'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2C3E50')
        cell.alignment = Alignment(horizontal='center')

    total = 0
    for idx, row in enumerate(sales_data, 1):
        date_str = row['date'].strftime('%d %b %Y') if hasattr(row['date'], 'strftime') else str(row['date'])[:10]
        ws.append([idx, row['customer'], date_str, row['total_amount'],
                   row['payment_status'], row['delivery_status']])
        total += row['total_amount']
        if idx % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill('solid', fgColor='F8F9FA')

    ws.append([])
    ws.append(['', 'TOTAL', '', total, '', ''])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col, w in zip('ABCDEF', [5, 25, 15, 14, 16, 18]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Customer Report ───────────────────────────────────────────────────────────

def generate_customer_report_pdf(customer_data, period_label):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors

    buf = io.BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    W, H = A4

    _header(pdf, f'Customer Report — {period_label}', f'Total Customers: {len(customer_data)}')

    y = H - 110
    cols = [40, 80, 220, 310, 380, 450, 510]
    headers = ['#', 'Shop Name', 'Owner', 'Phone', 'Village', 'Orders', 'Spent (₹)']
    pdf.setFillColor(colors.HexColor('#2c3e50'))
    pdf.rect(30, y - 5, W - 60, 20, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 8)
    for i, h in enumerate(headers):
        pdf.drawString(cols[i], y + 3, h)

    pdf.setFont('Helvetica', 8)
    y -= 20
    for idx, row in enumerate(customer_data):
        if y < 60:
            _footer(pdf)
            pdf.showPage()
            y = H - 50
        bg = colors.HexColor('#f8f9fa') if idx % 2 == 0 else colors.white
        pdf.setFillColor(bg)
        pdf.rect(30, y - 5, W - 60, 18, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor('#2c3e50'))
        vals = [str(idx + 1), str(row['shop_name'])[:18], str(row['owner_name'])[:12],
                str(row['phone']), str(row['village'])[:10],
                str(row['total_orders']), f"₹{row['total_spent']:,.0f}"]
        for i, v in enumerate(vals):
            pdf.drawString(cols[i], y + 2, v)
        y -= 18

    _footer(pdf)
    pdf.save()
    buf.seek(0)
    return buf


def generate_customer_report_excel(customer_data, period_label):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Customer Report'

    ws.merge_cells('A1:G1')
    ws['A1'] = f'ColdSync Pro — Customer Report — {period_label}'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='C0392B')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = f'Generated: {timezone.now().strftime("%d %b %Y %I:%M %p")}'
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])
    ws.append(['#', 'Shop Name', 'Owner Name', 'Phone', 'Village', 'Total Orders', 'Total Spent (₹)'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2C3E50')
        cell.alignment = Alignment(horizontal='center')

    for idx, row in enumerate(customer_data, 1):
        ws.append([idx, row['shop_name'], row['owner_name'], row['phone'],
                   row['village'], row['total_orders'], row['total_spent']])
        if idx % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill('solid', fgColor='F8F9FA')

    for col, w in zip('ABCDEFG', [5, 25, 20, 15, 15, 14, 16]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Products Report ───────────────────────────────────────────────────────────

def generate_products_report_pdf(products_data, period_label):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors

    buf = io.BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    W, H = A4

    _header(pdf, f'Top Products Report — {period_label}', f'Total Products: {len(products_data)}')

    y = H - 110
    cols = [40, 80, 210, 290, 360, 430, 500]
    headers = ['#', 'Product', 'Brand', 'Size', 'Bottles', 'Orders', 'Revenue (₹)']
    pdf.setFillColor(colors.HexColor('#2c3e50'))
    pdf.rect(30, y - 5, W - 60, 20, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 8)
    for i, h in enumerate(headers):
        pdf.drawString(cols[i], y + 3, h)

    pdf.setFont('Helvetica', 8)
    y -= 20
    for idx, row in enumerate(products_data):
        if y < 60:
            _footer(pdf)
            pdf.showPage()
            y = H - 50
        bg = colors.HexColor('#f8f9fa') if idx % 2 == 0 else colors.white
        pdf.setFillColor(bg)
        pdf.rect(30, y - 5, W - 60, 18, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor('#2c3e50'))
        vals = [str(idx + 1), str(row['product_name'])[:18], str(row['brand'])[:12],
                str(row['bottle_size']), str(row['total_bottles']),
                str(row['order_count']), f"₹{row['total_revenue']:,.0f}"]
        for i, v in enumerate(vals):
            pdf.drawString(cols[i], y + 2, v)
        y -= 18

    _footer(pdf)
    pdf.save()
    buf.seek(0)
    return buf


def generate_products_report_excel(products_data, period_label):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products Report'

    ws.merge_cells('A1:G1')
    ws['A1'] = f'ColdSync Pro — Top Products Report — {period_label}'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='C0392B')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = f'Generated: {timezone.now().strftime("%d %b %Y %I:%M %p")}'
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])
    ws.append(['#', 'Product Name', 'Brand', 'Bottle Size', 'Total Bottles', 'Order Count', 'Total Revenue (₹)'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2C3E50')
        cell.alignment = Alignment(horizontal='center')

    for idx, row in enumerate(products_data, 1):
        ws.append([idx, row['product_name'], row['brand'], row['bottle_size'],
                   row['total_bottles'], row['order_count'], row['total_revenue']])
        if idx % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill('solid', fgColor='F8F9FA')

    for col, w in zip('ABCDEFG', [5, 25, 18, 12, 14, 12, 18]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Monthly Revenue Report ────────────────────────────────────────────────────

def generate_monthly_report_pdf(months_data):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors

    buf = io.BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    total = sum(m['revenue'] for m in months_data)

    _header(pdf, 'Monthly Revenue Report', f'Total Revenue: ₹{total:,.2f}')

    y = H - 110
    cols = [40, 160, 280, 380, 460]
    headers = ['Month', 'Revenue (₹)', 'Orders', 'Avg Order (₹)', 'Paid (₹)']
    pdf.setFillColor(colors.HexColor('#2c3e50'))
    pdf.rect(30, y - 5, W - 60, 20, fill=True, stroke=False)
    pdf.setFillColor(colors.white)
    pdf.setFont('Helvetica-Bold', 10)
    for i, h in enumerate(headers):
        pdf.drawString(cols[i], y + 3, h)

    pdf.setFont('Helvetica', 10)
    y -= 22
    for idx, row in enumerate(months_data):
        bg = colors.HexColor('#f8f9fa') if idx % 2 == 0 else colors.white
        pdf.setFillColor(bg)
        pdf.rect(30, y - 5, W - 60, 20, fill=True, stroke=False)
        pdf.setFillColor(colors.HexColor('#2c3e50'))
        avg = row.get('avg_order', round(row['revenue'] / row['order_count'], 2) if row['order_count'] else 0)
        paid = row.get('paid', 0)
        vals = [row['month'], f"₹{row['revenue']:,.0f}", str(row['order_count']),
                f"₹{avg:,.0f}", f"₹{paid:,.0f}"]
        for i, v in enumerate(vals):
            pdf.drawString(cols[i], y + 3, v)
        y -= 22

    _footer(pdf)
    pdf.save()
    buf.seek(0)
    return buf


def generate_monthly_report_excel(months_data):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Monthly Revenue'

    ws.merge_cells('A1:E1')
    ws['A1'] = 'ColdSync Pro — Monthly Revenue Report'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='C0392B')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = f'Generated: {timezone.now().strftime("%d %b %Y %I:%M %p")}'
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])
    ws.append(['Month', 'Revenue (₹)', 'Order Count', 'Avg Order (₹)', 'Paid (₹)'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2C3E50')
        cell.alignment = Alignment(horizontal='center')

    total_rev = 0
    for idx, row in enumerate(months_data, 1):
        avg = row.get('avg_order', round(row['revenue'] / row['order_count'], 2) if row['order_count'] else 0)
        paid = row.get('paid', 0)
        ws.append([row['month'], row['revenue'], row['order_count'], avg, paid])
        total_rev += row['revenue']
        if idx % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill('solid', fgColor='F8F9FA')

    ws.append([])
    ws.append(['TOTAL', total_rev, '', '', ''])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col, w in zip('ABCDE', [15, 16, 14, 16, 14]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
